import math, datetime
from openpyxl import Workbook


# Workbook = 엑셀 파일 
wb = Workbook()
print(wb.worksheets)
print(wb.worksheets[0])
print(wb['Sheet'])   # <Worksheet "Sheet">
print(wb.sheetnames) # ['Sheet']


# worksheet 제목 바꾸기 
ws = wb.worksheets[0] # 엑셀 파일 속 하나의 tab = worksheet 
ws.title = '주소록'
print(ws.title)

ws['A1'] = '이름'
ws['B1'] = '전화번호'
ws['A2'] = '홍길동'
ws['B2'] = '777'


# Cell 단위로 접근할 수 있음 
cell = ws.cell(ros=3, column=1)
cell.value = '강동원'
ws.cesll(row=3, column=2).value = 147


# 새로운 워크시트(탭)을 추가하여 저장할 수 있음 
wb.create_sheet['new one']
print(wb.sheetnames)
ws = wb['new one']
ws['A1'] = '이름'
ws['A2'] = 12345
ws['A3'] = math.pi
ws['A4'] = datetime.datetime(2023, 11, 22, 10, 0, 0)
ws['A5'] = '=SIN(PI()/2)'
wb.save('address.xlsx') # xlsx는 zip파일임 